#!/usr/bin/env python3
"""
Script to export hierarchy.json to XLSX format with nama_unit, nama_parent, eselon, and jabatan columns.

Usage:
    python export_to_xlsx.py
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: openpyxl library is required. Install it with: pip install openpyxl")
    sys.exit(1)


def simplify_jabatan(jabatan):
    """
    Simplify jabatan by removing unit-specific names.
    
    Args:
        jabatan: Full position title
    
    Returns:
        Simplified position title
    """
    if not jabatan:
        return jabatan
    
    jabatan_lower = jabatan.lower()
    
    # Rules for simplification:
    # 1. "Kepala Badan X" -> "Kepala Badan"
    if jabatan_lower.startswith('kepala badan '):
        return 'Kepala Badan'
    
    # 2. "Sekretaris Badan X" -> "Sekretaris Badan"
    elif jabatan_lower.startswith('sekretaris badan '):
        return 'Sekretaris Badan'
    
    # 3. "Kepala Dinas X" -> "Kepala Dinas"
    elif jabatan_lower.startswith('kepala dinas '):
        return 'Kepala Dinas'
    
    # 4. "Sekretaris Dinas X" -> "Sekretaris Dinas" 
    elif jabatan_lower.startswith('sekretaris dinas '):
        return 'Sekretaris Dinas'
    
    # 5. "Direktur X" -> "Direktur"
    elif jabatan_lower.startswith('direktur '):
        return 'Direktur'
    
    # 6. "Kepala Bidang X" -> "Kepala Bidang"
    elif jabatan_lower.startswith('kepala bidang '):
        return 'Kepala Bidang'
    
    # 7. "Kepala Bagian X" -> "Kepala Bagian"
    elif jabatan_lower.startswith('kepala bagian '):
        return 'Kepala Bagian'
    
    # 8. "Kepala Subbagian X" -> "Kepala Subbagian"
    elif jabatan_lower.startswith('kepala subbagian '):
        return 'Kepala Subbagian'
    
    # 9. "Kepala Sub Bagian X" -> "Kepala Sub Bagian"
    elif jabatan_lower.startswith('kepala sub bagian '):
        return 'Kepala Sub Bagian'
    
    # 10. "Kepala Subbidang X" -> "Kepala Subbidang"
    elif jabatan_lower.startswith('kepala subbidang '):
        return 'Kepala Subbidang'
    
    # 11. "Kepala Sub Bidang X" -> "Kepala Sub Bidang"
    elif jabatan_lower.startswith('kepala sub bidang '):
        return 'Kepala Sub Bidang'
    
    # 12. "Kepala Seksi X" -> "Kepala Seksi"
    elif jabatan_lower.startswith('kepala seksi '):
        return 'Kepala Seksi'
    
    # 13. "Wakil Direktur X" -> "Wakil Direktur"
    elif jabatan_lower.startswith('wakil direktur '):
        return 'Wakil Direktur'
    
    # 14. "Asisten Sekretaris Daerah" -> "Asisten"
    elif jabatan_lower == 'asisten sekretaris daerah':
        return 'Asisten'
    
    # 15. "Inspektur Pembantu I/II/III/IV/V" -> "Inspektur Pembantu"
    elif jabatan_lower.startswith('inspektur pembantu '):
        return 'Inspektur Pembantu'
    
    # 16. "Kepala UPKP" (special case, should remain as is)
    # 17. "Kepala UPTD" (special case, should remain as is)
    # 18. "Kepala Unit" for UPTD (should remain as is)
    # For other cases, return as is
    return jabatan


def generate_kode_jabatan(jabatan, unit_name):
    """
    Generate a position code (kode_jabatan) based on the position title and unit name.
    
    Args:
        jabatan: Position title
        unit_name: Unit name
    
    Returns:
        Position code string
    """
    jabatan_lower = jabatan.lower()
    unit_lower = unit_name.lower()
    
    # Sekretaris Daerah
    if "sekretaris daerah" in jabatan_lower:
        return "SEKDA"
    
    # Sekretaris DPRD
    if "sekretaris dprd" in jabatan_lower:
        return "SEKDPRD"
    
    # Kepala Pelaksana BPBD
    if "kepala pelaksana bpbd" in jabatan_lower:
        return "KAPEL_BPBD"
    
    # Kepala Badan
    if "kepala badan" in jabatan_lower:
        # Extract key words from unit name
        if "kepegawaian" in unit_lower:
            return "KABKP"
        elif "keuangan" in unit_lower:
            return "KABKAD"
        elif "penanggulangan bencana" in unit_lower or "bpbd" in unit_lower:
            return "KABPBD"
        elif "kesatuan bangsa" in unit_lower:
            return "KABKESBANGPOL"
        elif "pendapatan" in unit_lower:
            return "KABAPENDA"
        elif "perencanaan" in unit_lower:
            return "KABAPPEDA"
        else:
            return "KABADAN"
    
    # Kepala Dinas
    if "kepala dinas" in jabatan_lower:
        if "pendidikan" in unit_lower:
            return "KADIS_DIKDAS"
        elif "kesehatan" in unit_lower:
            return "KADIS_KES"
        elif "pekerjaan umum" in unit_lower or "perumahan" in unit_lower:
            return "KADIS_PUPR"
        elif "sosial" in unit_lower:
            return "KADIS_SOSIAL"
        elif "lingkungan" in unit_lower:
            return "KADIS_LH"
        else:
            return "KADIS"
    
    # Kepala Bagian
    if "kepala bagian" in jabatan_lower:
        if "umum" in unit_lower and "protokol" in unit_lower:
            return "KB_UMUM_PROT"
        elif "umum" in unit_lower:
            return "KB_UMUM"
        elif "keuangan" in unit_lower:
            return "KB_KEUANGAN"
        else:
            return "KABAG"
    
    # Kepala Bidang
    if "kepala bidang" in jabatan_lower:
        if "pendidikan dasar" in unit_lower:
            return "KABID_DIKDAS"
        else:
            return "KABID"
    
    # Kepala Subbagian
    if "kepala subbagian" in jabatan_lower or "kepala sub bagian" in jabatan_lower:
        if "keuangan" in unit_lower:
            return "KASUBAG_KEU"
        elif "umum" in unit_lower:
            return "KASUBAG_UMUM"
        else:
            return "KASUBAG"
    
    # Sekretaris (for agencies/organizations)
    if "sekretaris" in jabatan_lower and "sekretaris daerah" not in jabatan_lower:
        return "SEKRET"
    
    # Inspektur
    if "inspektur" in jabatan_lower:
        return "INSPEKTUR"
    
    # Camat
    if "camat" in jabatan_lower:
        return "CAMAT"
    
    # Default
    return "KODE"


def flatten_hierarchy(data, parent_name=""):
    """
    Flatten hierarchical JSON structure into a list of dictionaries with all required fields.
    
    Args:
        data: List of organizational units with nested children
        parent_name: Name of the parent unit (empty string for top-level)
    
    Returns:
        List of dictionaries with keys: nama_unit, nama_parent, eselon, jabatan, jabatan_lengkap, kode_jabatan, catatan
    """
    result = []
    
    if isinstance(data, list):
        for item in data:
            if isinstance(item, dict) and 'name' in item:
                unit_name = item['name']
                eselon = item.get('eselon', '')
                jabatan_original = item.get('jabatan', '')
                
                # Special handling for units where jabatan is "Kepala" but needs more context
                if jabatan_original == 'Kepala':
                    unit_lower = unit_name.lower()
                    
                    # UPTD units should have "Kepala Unit" as jabatan
                    if unit_name.startswith('UPTD '):
                        jabatan_lengkap = 'Kepala Unit'
                        jabatan_original = 'Kepala Unit'
                    
                    # Inspektur Pembantu I-V should have full title
                    elif 'inspektur pembantu' in unit_lower:
                        jabatan_lengkap = unit_name  # e.g., "Inspektur Pembantu I"
                        jabatan_original = unit_name
                    
                    else:
                        jabatan_lengkap = jabatan_original
                else:
                    jabatan_lengkap = jabatan_original
                
                # Generate additional fields
                jabatan = simplify_jabatan(jabatan_original)  # Simplify for jabatan column
                kode_jabatan = generate_kode_jabatan(jabatan_original, unit_name)
                catatan = item.get('catatan', '')  # Get catatan from JSON if exists, otherwise empty
                
                record = {
                    'nama_unit': unit_name,
                    'nama_parent': parent_name,
                    'eselon': eselon,
                    'jabatan': jabatan,
                    'jabatan_lengkap': jabatan_lengkap,
                    'kode_jabatan': kode_jabatan,
                    'catatan': catatan
                }
                result.append(record)
                
                # Recursively process children
                if 'children' in item and isinstance(item['children'], list):
                    children_results = flatten_hierarchy(item['children'], unit_name)
                    result.extend(children_results)
    
    return result


def create_xlsx(data, output_file="hierarchy_export.xlsx"):
    """
    Create XLSX file with all required columns.
    
    Args:
        data: List of dictionaries with keys: nama_unit, nama_parent, eselon, jabatan, jabatan_lengkap, kode_jabatan, catatan
        output_file: Path to output XLSX file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Hierarchy"
    
    # Set column headers
    ws['A1'] = "nama_unit"
    ws['B1'] = "nama_parent"
    ws['C1'] = "eselon"
    ws['D1'] = "jabatan"
    ws['E1'] = "jabatan_lengkap"
    ws['F1'] = "kode_jabatan"
    ws['G1'] = "catatan"
    
    # Style headers
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = header_alignment
    
    # Write data
    for idx, record in enumerate(data, start=2):
        ws[f'A{idx}'] = record.get('nama_unit', '')
        ws[f'B{idx}'] = record.get('nama_parent', '')
        ws[f'C{idx}'] = record.get('eselon', '')
        ws[f'D{idx}'] = record.get('jabatan', '')
        ws[f'E{idx}'] = record.get('jabatan_lengkap', '')
        ws[f'F{idx}'] = record.get('kode_jabatan', '')
        ws[f'G{idx}'] = record.get('catatan', '')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 30
    
    # Save workbook
    wb.save(output_file)
    print(f"Successfully created {output_file}")
    print(f"Total records: {len(data)}")


def main():
    """Main function to export hierarchy to XLSX."""
    # Read hierarchy.json
    json_file = Path(__file__).parent / "hierarchy.json"
    
    if not json_file.exists():
        print(f"Error: {json_file} not found!")
        sys.exit(1)
    
    print(f"Reading {json_file}...")
    with open(json_file, 'r', encoding='utf-8') as f:
        hierarchy_data = json.load(f)
    
    print("Flattening hierarchy...")
    flattened_data = flatten_hierarchy(hierarchy_data)
    
    # Create XLSX file
    output_file = Path(__file__).parent / "hierarchy_export.xlsx"
    print(f"Creating {output_file}...")
    create_xlsx(flattened_data, str(output_file))
    
    print("\nFirst 5 records:")
    for i, record in enumerate(flattened_data[:5], 1):
        print(f"{i}. Unit: {record['nama_unit']}")
        print(f"   Parent: {record['nama_parent'] if record['nama_parent'] else '(kosong)'}")
        print(f"   Eselon: {record['eselon'] if record['eselon'] else '(kosong)'}")
        print(f"   Jabatan: {record['jabatan'] if record['jabatan'] else '(kosong)'}")
        print(f"   Jabatan Lengkap: {record['jabatan_lengkap'] if record['jabatan_lengkap'] else '(kosong)'}")
        print(f"   Kode Jabatan: {record['kode_jabatan'] if record['kode_jabatan'] else '(kosong)'}")
        print(f"   Catatan: {record['catatan'] if record['catatan'] else '(kosong)'}")
        print()


if __name__ == "__main__":
    main()
